"""导入和远智能客户材料包到数据库与 MinIO。"""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import SessionLocal, engine, Base
from app.models import (
    CompanyProfile,
    Template,
    Qualification,
    ChecklistItem,
    FAQItem,
    FieldDef,
)
from app.services import storage
from app.services.word import extract_placeholders
from app.seed.load_sample import wait_deps

settings = get_settings()

MIME = {
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def pack_root() -> Path:
    env = Path(getattr(settings, "customer_data_dir", "") or "")
    if env and env.exists():
        return env
    root = Path(__file__).resolve().parents[3]
    cand = root / "customer_data" / "heyuanzhineng_20260729"
    return cand


def read_xlsx_sheet(path: Path, sheet: str | None = None) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])]
    out = []
    for row in rows[1:]:
        if not any(row):
            continue
        out.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return out


def parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s in ("无", "长期", "None"):
        return None
    try:
        return date.fromisoformat(s[:10])
    except Exception:
        return None


def clear_base_data(db: Session):
    from app.models import TenderProject, ProjectSnapshot, ProjectExport
    # 解除项目对模板的外键引用（强制重导会清空演示项目）
    db.query(ProjectExport).delete()
    db.query(ProjectSnapshot).delete()
    db.query(TenderProject).delete()
    db.query(Qualification).delete()
    db.query(ChecklistItem).delete()
    db.query(FAQItem).delete()
    db.query(FieldDef).delete()
    db.query(Template).delete()
    db.query(CompanyProfile).delete()
    db.commit()


def import_company(db: Session, base: Path):
    info = (base / "pack" / "01-企业与项目基础信息" / "1.企业基本信息【必填】.txt").read_text(encoding="utf-8", errors="ignore")
    typical = ""
    tp = base / "pack" / "01-企业与项目基础信息" / "2.典型投标项目说明【建议】.txt"
    if tp.exists():
        typical = tp.read_text(encoding="utf-8", errors="ignore")
    style = ""
    sp = base / "pack" / "03-历史投标文件" / "3-AI生成内容风格参考【建议】.txt"
    if sp.exists():
        style = sp.read_text(encoding="utf-8", errors="ignore")

    def pick(label: str) -> str:
        m = re.search(rf"{label}[：:]\s*(.+)", info)
        return (m.group(1).strip() if m else "")

    # 简介多行
    intro_m = re.search(r"企业简介[^：:]*[：:]\s*\n?(.*?)(?:\n\s*主营业务|$)", info, re.S)
    intro = (intro_m.group(1).strip() if intro_m else "")
    biz_m = re.search(r"主营业务[：:]\s*(.+)", info)
    qual_m = re.search(r"铁路/轨道交通相关资质概述[：:]\s*\n?(.*)", info, re.S)

    profile = CompanyProfile(
        id=1,
        full_name=pick("企业全称") or "和远智能科技股份有限公司",
        short_name=pick("企业简称") or "和远智能",
        credit_code=pick("统一社会信用代码"),
        register_address=pick("注册地址"),
        office_address=pick("办公地址"),
        legal_name=pick("法定代表人"),
        legal_gender="男",
        legal_age="60",
        legal_title="董事长",
        legal_id_no="3700000000",
        registered_capital=pick("注册资本"),
        founded_date=pick("成立日期"),
        phone=pick("联系电话"),
        fax="0531-68621679",
        email=pick("企业邮箱"),
        website="www.hyzn77.com",
        postcode="250100",
        bank_name=pick("开户银行"),
        bank_account=pick("银行账号"),
        recent_revenue="2023：****万 2024：****万 2025：****万",
        related_companies="无",
        intro=intro,
        business_scope=biz_m.group(1).strip() if biz_m else "",
        qual_overview=(qual_m.group(1).strip() if qual_m else ""),
        typical_projects=typical,
        ai_style_notes=style,
    )
    db.merge(profile)
    db.commit()
    print("[import] company ok")


def import_fields(db: Session, base: Path):
    meta = json.loads((base / "field_keys.json").read_text(encoding="utf-8"))
    company = db.query(CompanyProfile).filter(CompanyProfile.id == 1).first()
    company_map = {}
    if company:
        company_map = {c.key: getattr(company, c.key, "") for c in CompanyProfile.__table__.columns if c.key != "id"}

    for i, f in enumerate(meta.get("fields") or []):
        default = f.get("default_value") or ""
        cf = f.get("company_field") or ""
        if f.get("is_company_default") and cf and company_map.get(cf):
            default = str(company_map.get(cf) or default)
        db.add(FieldDef(
            name=f["name"],
            key=f["key"],
            field_type=f.get("field_type") or "文本",
            required=bool(f.get("required", True)),
            default_value=default,
            options=f.get("options") or "",
            module=f.get("module") or "",
            validation="",
            template_code=f.get("template_code") or "common",
            sort_order=i,
            is_company_default=bool(f.get("is_company_default")),
            company_field=cf,
            desensitized=bool(f.get("desensitized")),
        ))
    db.commit()
    print("[import] fields ok")


def import_templates(db: Session, base: Path):
    eng = base / "engineered_templates"
    for f in sorted(eng.glob("*.docx")):
        if f.name.startswith("engineer"):
            continue
        code = "common"
        kind = "template"
        if "模板1" in f.name:
            code = "tpl1"
        elif "模板3" in f.name:
            code = "tpl3"
        elif "骨架" in f.name:
            code = "common"
            kind = "skeleton"
        key = f"templates/{f.name}"
        storage.upload_file(key, str(f), MIME[".docx"])
        data = f.read_bytes()
        ph = extract_placeholders(data)
        db.add(Template(
            name=f.stem,
            description="工程化可替换模板",
            object_key=key,
            placeholders={"list": ph},
            template_code=code,
            kind=kind,
            enabled=True,
        ))

    hist_dir = base / "pack" / "03-历史投标文件" / "1-完整历史标书"
    if hist_dir.exists():
        for f in sorted(hist_dir.glob("*.docx")):
            is_tender = "招标文件" in f.name
            key = f"history/{f.name}"
            storage.upload_file(key, str(f), MIME[".docx"])
            # 历史快照：从文件名粗提取
            snap = {"project_name": f.stem, "bidder_name": "和远智能科技股份有限公司"}
            db.add(Template(
                name=f.stem,
                description="招标文件（只读参考）" if is_tender else "历史标书（智能替换起点）",
                object_key=key,
                placeholders={"list": [], "is_history": not is_tender},
                template_code="history",
                kind="tender_doc" if is_tender else "history",
                enabled=not is_tender,
                source_snapshot=snap,
            ))
    db.commit()
    print("[import] templates ok")


def find_qual_file(qual_root: Path, fname: str) -> Path | None:
    if not fname:
        return None
    direct = list(qual_root.rglob(fname))
    if direct:
        return direct[0]
    stem = Path(fname).stem.lower()
    for p in qual_root.rglob("*"):
        if not p.is_file():
            continue
        if stem and stem in p.name.lower():
            return p
        # 清单名 vs 磁盘名轻微差异
        if stem[:12] and stem[:12] in p.name.lower():
            return p
    return None


SECTION_HINTS = {
    "企业资质包": "基本情况表",
    "业绩包": "业绩",
    "人员信息包": "人员信息",
    "财务信息包": "财务信息",
    "信誉与法律包": "信誉情况",
    "技术方案包": "质量保证能力证明",
    "商务文件包": "商务文件",
}


def import_qualifications(db: Session, base: Path):
    qual_root = base / "pack" / "06-资质数据库参考资料"
    list_file = next(qual_root.glob("*资质材料清单*.xlsx"), None)
    if not list_file:
        print("[import] qualifications list missing")
        return
    rows = read_xlsx_sheet(list_file)
    for i, row in enumerate(rows):
        fname = str(row.get("文件名") or "").strip()
        category = str(row.get("分类") or "").strip()
        found = find_qual_file(qual_root, fname)
        if not found:
            # try material name
            continue
        ext = found.suffix.lower()
        key = f"qualifications/{category}/{found.name}"
        storage.upload_file(key, str(found), MIME.get(ext, "application/octet-stream"))
        long_term = str(row.get("是否长期有效") or "") == "是"
        db.add(Qualification(
            category=category,
            name=str(row.get("材料名称") or found.stem),
            issuer=str(row.get("颁发机构") or ""),
            file_type=str(row.get("文件类型") or ext.lstrip(".")),
            file_name=found.name,
            object_key=key,
            keywords=str(row.get("关键词标签") or ""),
            section_hint=SECTION_HINTS.get(category, ""),
            sort_order=i,
            valid_from=parse_date(row.get("有效期起")),
            valid_to=parse_date(row.get("有效期止")),
            is_long_term=long_term,
            ocr_text=f"{row.get('材料名称')} {row.get('颁发机构')} {row.get('关键词标签')}",
        ))
    db.commit()
    print("[import] qualifications ok")


def import_checklist(db: Session, base: Path):
    mapping = {
        "1-商务标条目清单.xlsx": "商务",
        "2-技术标条目清单.xlsx": "技术",
        "3-报价标条目清单.xlsx": "报价",
        "4-资质材料必附清单.xlsx": "资质",
    }
    folder = base / "pack" / "05-条目完整性校验清单"
    order = 0
    for fname, section in mapping.items():
        path = folder / fname
        if not path.exists():
            continue
        wb = load_workbook(path, data_only=True)
        for sheet in wb.sheetnames:
            code = "common"
            if "模板1" in sheet:
                code = "tpl1"
            elif "模板3" in sheet:
                code = "tpl3"
            for row in read_xlsx_sheet(path, sheet):
                name = str(row.get("条目名称") or row.get("资质名称") or "").strip()
                if not name:
                    continue
                db.add(ChecklistItem(
                    section=section,
                    name=name,
                    required=str(row.get("是否必含") or row.get("是否必附") or "必含"),
                    chapter=str(row.get("对应章节") or row.get("所属分类") or ""),
                    remark=str(row.get("备注") or ""),
                    template_code=code,
                    sort_order=order,
                ))
                order += 1
    db.commit()
    print("[import] checklist ok")


def import_faq(db: Session, base: Path):
    path = next((base / "pack" / "07-企业问答chatbot参考").glob("*.xlsx"), None)
    if not path:
        return
    for row in read_xlsx_sheet(path):
        q = str(row.get("问题内容") or "").strip()
        if not q:
            continue
        db.add(FAQItem(
            category=str(row.get("问题类别") or ""),
            question=q,
            answer=str(row.get("标准答案") or ""),
            source=str(row.get("答案来源材料") or ""),
            template_code="common",
        ))
    db.commit()
    print("[import] faq ok")


def run_import(force: bool = False) -> dict:
    wait_deps()
    from app.database_migrate import ensure_schema
    ensure_schema()
    base = pack_root()
    if not (base / "field_keys.json").exists():
        raise RuntimeError(f"客户包不存在或未准备: {base}")
    if not (base / "engineered_templates").exists():
        raise RuntimeError("请先运行 scripts/engineer_templates.py")

    db = SessionLocal()
    try:
        has_data = db.query(FieldDef).count() > 0
        if has_data and not force:
            print("[import] 已有基础数据，跳过（使用 force=True 覆盖）")
            return {"skipped": True, "root": str(base)}
        if force:
            clear_base_data(db)
        import_company(db, base)
        import_fields(db, base)
        import_templates(db, base)
        import_qualifications(db, base)
        import_checklist(db, base)
        import_faq(db, base)
        return {"ok": True, "root": str(base)}
    finally:
        db.close()


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--force", action="store_true")
    args = p.parse_args()
    print(run_import(force=args.force))
