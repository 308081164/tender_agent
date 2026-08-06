"""从 sample_data 初始化数据库与 MinIO"""
from __future__ import annotations
import time
from pathlib import Path
from datetime import date

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import SessionLocal, engine, Base
from app.models import (
    Template, Qualification, ChecklistItem, FAQItem, FieldDef,
)
from app.services import storage
from app.services.word import extract_placeholders

settings = get_settings()


def wait_deps(max_retries: int = 30):
    for i in range(max_retries):
        try:
            with engine.connect() as conn:
                conn.execute(__import__("sqlalchemy").text("SELECT 1"))
            storage.ensure_bucket()
            return
        except Exception as e:
            print(f"[seed] waiting deps... ({i+1}/{max_retries}) {e}")
            time.sleep(2)
    raise RuntimeError("Dependencies not ready")


def read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return result


def seed_templates(db: Session, root: Path):
    if db.query(Template).count() > 0:
        return
    tpl_dir = root / "02_Word标书模板" / "标准投标Word模板"
    for f in tpl_dir.glob("*.docx"):
        key = f"templates/{f.name}"
        storage.upload_file(key, str(f), "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        data = f.read_bytes()
        placeholders = extract_placeholders(data)
        db.add(Template(
            name=f.stem,
            description="铁路工程投标通用模板（模拟）",
            object_key=key,
            placeholders={"list": placeholders},
        ))
    # 历史标书也作为可选起点
    hist_dir = root / "03_历史投标文件" / "完整历史标书"
    for f in hist_dir.glob("*.docx"):
        key = f"history/{f.name}"
        storage.upload_file(key, str(f), "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        data = f.read_bytes()
        placeholders = extract_placeholders(data)
        db.add(Template(
            name=f.stem,
            description="历史标书（可复用）",
            object_key=key,
            placeholders={"list": placeholders, "is_history": True},
        ))
    db.commit()
    print("[seed] templates ok")


def seed_qualifications(db: Session, root: Path):
    if db.query(Qualification).count() > 0:
        return
    qual_root = root / "06_资质数据库参考材料"
    list_file = qual_root / "资质材料清单.xlsx"
    rows = read_xlsx(list_file) if list_file.exists() else []
    for row in rows:
        fname = str(row.get("文件名") or "")
        category = str(row.get("分类") or "")
        # find file
        found = None
        for p in qual_root.rglob(fname):
            found = p
            break
        if not found:
            continue
        key = f"qualifications/{category}/{fname}"
        storage.upload_file(key, str(found), "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        valid_to = None
        valid_from = None
        try:
            vt = row.get("有效期止")
            if vt and str(vt) != "长期":
                valid_to = date.fromisoformat(str(vt)[:10])
            vf = row.get("有效期起")
            if vf:
                valid_from = date.fromisoformat(str(vf)[:10])
        except Exception:
            pass
        long_term = str(row.get("是否长期有效") or "") == "是"
        ocr = f"{row.get('材料名称')} {row.get('颁发机构')} {row.get('关键词标签')}"
        db.add(Qualification(
            category=category,
            name=str(row.get("材料名称") or fname),
            issuer=str(row.get("颁发机构") or ""),
            file_type=str(row.get("文件类型") or "docx"),
            object_key=key,
            keywords=str(row.get("关键词标签") or ""),
            valid_from=valid_from,
            valid_to=valid_to,
            is_long_term=long_term,
            ocr_text=ocr,
        ))
    db.commit()
    print("[seed] qualifications ok")


def seed_checklist(db: Session, root: Path):
    if db.query(ChecklistItem).count() > 0:
        return
    mapping = {
        "商务标条目清单.xlsx": "商务",
        "技术标条目清单.xlsx": "技术",
        "报价标条目清单.xlsx": "报价",
        "资质材料必附清单.xlsx": "资质",
    }
    base = root / "05_条目完整性校验清单"
    for fname, section in mapping.items():
        path = base / fname
        if not path.exists():
            continue
        for row in read_xlsx(path):
            name = str(row.get("条目名称") or row.get("资质名称") or "")
            if not name:
                continue
            db.add(ChecklistItem(
                section=section,
                name=name,
                required=str(row.get("是否必含") or row.get("是否必附") or "必含"),
                chapter=str(row.get("对应章节") or row.get("所属分类") or ""),
                remark=str(row.get("备注") or ""),
            ))
    db.commit()
    print("[seed] checklist ok")


def seed_faq(db: Session, root: Path):
    if db.query(FAQItem).count() > 0:
        return
    path = root / "07_企业问答Chatbot参考" / "招标常见问题与标准答案.xlsx"
    for row in read_xlsx(path):
        db.add(FAQItem(
            category=str(row.get("问题类别") or ""),
            question=str(row.get("问题内容") or ""),
            answer=str(row.get("标准答案") or ""),
            source=str(row.get("答案来源材料") or ""),
        ))
    db.commit()
    print("[seed] faq ok")


def seed_fields(db: Session, root: Path):
    if db.query(FieldDef).count() > 0:
        return
    fields_path = root / "04_结构化字段与录入规范" / "标书关键字段清单.xlsx"
    options_path = root / "04_结构化字段与录入规范" / "下拉选项枚举值.xlsx"
    options_map = {}
    if options_path.exists():
        for row in read_xlsx(options_path):
            options_map[str(row.get("字段名称") or "")] = str(
                row.get("可选值（多个值用中文分号分隔）") or row.get("可选值") or ""
            )
    for row in read_xlsx(fields_path):
        name = str(row.get("字段名称") or "")
        key = str(row.get("字段英文名") or "")
        if not key:
            continue
        db.add(FieldDef(
            name=name,
            key=key,
            field_type=str(row.get("字段类型") or "文本"),
            required=str(row.get("是否必填") or "是") == "是",
            default_value=str(row.get("默认值") or ""),
            options=options_map.get(name, ""),
            module=str(row.get("所属模块") or ""),
            validation=str(row.get("校验规则") or ""),
        ))
    db.commit()
    print("[seed] fields ok")


def run_seed():
    wait_deps()
    Base.metadata.create_all(bind=engine)
    root = Path(settings.sample_data_dir)
    if not root.exists():
        # 本地开发回退
        root = Path(__file__).resolve().parents[3] / "sample_data"
    print(f"[seed] sample_data = {root}")
    db = SessionLocal()
    try:
        seed_templates(db, root)
        seed_qualifications(db, root)
        seed_checklist(db, root)
        seed_faq(db, root)
        seed_fields(db, root)
        print("[seed] done")
    finally:
        db.close()


if __name__ == "__main__":
    run_seed()
